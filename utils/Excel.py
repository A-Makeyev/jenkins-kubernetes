import openpyxl

def get_count(file, sheet_name, count_type):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row if count_type == 'row' else sheet.max_column if count_type == 'column' else None


def read_data(file, sheet_name, row_number, column_number):
    workbook = openpyxl.load_workbook(file)
    return workbook[sheet_name].cell(row=row_number, column=column_number).value


def write_data(file, sheet_name, row_number, column_number, data):
    workbook = openpyxl.load_workbook(file)
    workbook[sheet_name].cell(row=row_number, column=column_number).value = data
    workbook.save(file)
